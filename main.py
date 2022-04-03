from PyQt5 import QtWidgets
import sys
from PyQt5.QtWidgets import QFileDialog
from login import Ui_mainWindow
from openpyxl import Workbook,load_workbook
from PyQt5.QtCore import QObject,QThread,pyqtSignal
from secondPage import SecondPage
from genel import GeneralDetailPage
import os

class Worker(QObject):

    arbeiter_name_signal=pyqtSignal(list)
    arbeiterKasseGesamtSignal=pyqtSignal(list)
    finished = pyqtSignal()
    display = pyqtSignal(str)
    
    def __init__(self,line):
        super().__init__()
        self.line=line
        

    def getExcelValueWorker(self):
        wb = load_workbook(self.line)
        ws = wb.active
        arbeiter_name=[]
        arbeiterKasseGesamt=[]
        for i in range(21,ws.max_column):
            if str(ws.cell(3,i).value) != "None" and ws.cell(3,i).value !="MITARBEITER NAME":
                arbeiter_name.append(ws.cell(3,i).value)
                arbeiterKasseGesamt.append(i)
        self.arbeiter_name_signal.emit(arbeiter_name)
        self.arbeiterKasseGesamtSignal.emit(arbeiterKasseGesamt)
        self.finished.emit()

class Window(Ui_mainWindow,QtWidgets.QMainWindow):
    def __init__(self):
        super(Ui_mainWindow,self).__init__()
        self.setupUi(self)

        self.worker = 0
        self.arbeiter_name=[]
        self.arbeiterKasseGesamt=[]
        self.exelList=[]

        self.exelListDetected()
        
        # self.fileSearchButton.clicked.connect(self.browsefiles)

        self.readExcelFileButton.clicked.connect(self.excelRead)
        self.personalDetailButton.clicked.connect(self.personalDetail)
        self.generalDetailButton.clicked.connect(self.generelDetail)

    def exelListDetected(self):
        dosyalar = os.listdir(".")
        self.exelList= []
        for dosya in dosyalar:
            if dosya.endswith(".xlsx"):
                self.exelList.append(dosya)
        for i in range(len(self.exelList)):
            self.comboBox.addItem(self.exelList[i],i)

    # def browsefiles(self):
    #     fname=QFileDialog.getOpenFileName(self,'Open file','','EXCEL DOSYALARI (*.xlsx)')
    #     try:
    #         self.fileExtentionLine.setText(fname[0])
    #     except:
    #         self.fileExtentionLine.setText("DOSYA SEÇİLMEDİ")

    def excelRead(self):
        if self.comboBox.currentText=="":
            self.errorTextEdit.append("Dosya Bulunamadı.")
        else:
            self.errorTextEdit.append("İşlem Başlatıldı. Bekleyiniz... \n")
            self.getExcelValue()

    def getExcelValue(self):
        try:
            self.my_thread = QThread(parent=self)
            self.worker = Worker(self.comboBox.currentText())
            self.worker.moveToThread(self.my_thread)
            self.worker.display.connect(self.displayMain)
            self.worker.arbeiter_name_signal.connect(self.setArbeiterName)
            self.worker.arbeiterKasseGesamtSignal.connect(self.setArbeiterKasseGesamt)
            self.worker.finished.connect(self.displayPersonalList)
            self.my_thread.start()
            self.my_thread.started.connect(self.worker.getExcelValueWorker)
        except Exception as e:
            print(e)

    def setArbeiterName(self,workerList):
        self.arbeiter_name=workerList

    def displayMain(self,dizi):
        print(dizi)
    
    def setArbeiterKasseGesamt(self,workerList):
        self.arbeiterKasseGesamt=workerList

    def displayPersonalList(self):
        for i in range (0,len(self.arbeiter_name)):
            self.workerListComboBox.addItem(str(self.arbeiter_name[i]),i)
        self.errorTextEdit.append("İsimler Listelendi")
    
    def personalDetail(self):
        # print(self.workerListComboBox.currentText())w
        for i in range(len(self.arbeiter_name)):
            if self.arbeiter_name[i]==self.workerListComboBox.currentText():
                # print(self.arbeiterKasseGesamt[i])
                # self.personalDetailPage=SecondPage(self.ws,self.arbeiterKasseGesamt[i])
                self.personalDetailPage=SecondPage(self.comboBox.currentText(),self.arbeiterKasseGesamt[i],0)
                self.personalDetailPage.show()  
    
    def generelDetail(self):
        if self.comboBox.currentText()=="":
            self.errorTextEdit.setText("Dosya seçiniz...")
        else:
            self.generelDetailPage=GeneralDetailPage(self.comboBox.currentText())
            self.generelDetailPage.show()

if __name__ == "__main__":
        app = QtWidgets.QApplication(sys.argv)
        ui=Window()
        ui.show()
        sys.exit(app.exec_())
