from personalDataUpdate import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5 import QtWidgets
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from openpyxl import Workbook,load_workbook

class Worker(QObject):

   finished=pyqtSignal(str)

   def __init__(self,line,personalRow,name):
      super().__init__()
      self.line=line
      self.personalRow=personalRow
      self.name=name

   def nameUpdateFunctionWorker(self):
      wb = load_workbook(self.line)
      ws = wb.active
      c1 = ws.cell(row = 3, column = self.personalRow)
      c1.value=self.name
      wb.save(self.line)
      self.finished.emit("VERİ GÜNCELLENDİ.")

   def setTableWorker(self):
      wb = load_workbook(self.line)
      ws = wb.active
      c2 = ws.cell(row = 3, column = self.personalRow)
      c2.value=""
      for j in range(7,68,2):
         c1 = ws.cell(row = j, column = self.personalRow)
         cValueText1=0
         c1.value=int(cValueText1)
         c3 = ws.cell(row = j, column = self.personalRow+3)
         cValueText3=0
         c3.value=int(cValueText3)
      wb.save(self.line)
      self.finished.emit("VERİ GÜNCELLENDİ.")

class personalDatePAge(Ui_MainWindow,QtWidgets.QMainWindow):
   def __init__(self,line,row,personType):
      super().__init__()
      self.setupUi(self)
      self.line=line
      self.personalRow=row
      self.personType=0

      self.nameUpdate.clicked.connect(self.nameUpdateFunction)
      self.deleteAll.clicked.connect(self.setTable)
      

   def nameUpdateFunction(self):
      try:
         if self.nameLine.text()=="":
            self.messageBox("isim giriniz.")
         else:
            self.my_thread = QThread(parent=self)
            self.worker = Worker(self.line,self.personalRow,self.nameLine.text())
            self.worker.moveToThread(self.my_thread)
            self.worker.finished.connect(self.messageBox)
            self.my_thread.start()
            self.my_thread.started.connect(self.worker.nameUpdateFunctionWorker)
      except:
         pass

   def setTable(self):
      try:
         self.my_thread = QThread(parent=self)
         self.worker = Worker(self.line,self.personalRow,self.nameLine.text())
         self.worker.moveToThread(self.my_thread)
         self.worker.finished.connect(self.messageBox)
         self.my_thread.start()
         self.my_thread.started.connect(self.worker.setTableWorker)
      except:
         pass

   def messageBox(self,displayText):
      msgBox = QMessageBox()
      msgBox.about(self,"KİŞİ İŞLEMLERİ",displayText)
