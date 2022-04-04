from PyQt5 import QtWidgets
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtWidgets import *
from general import Ui_MainWindow
from openpyxl import Workbook,load_workbook
import pandas as pd


class Worker(QObject):

    row = pyqtSignal(int)
    col = pyqtSignal (int)
    value = pyqtSignal (str)
    writeSignal = pyqtSignal ()

    def __init__(self,line):
        super().__init__()
        self.line=line

    def setTable(self):

        wb = load_workbook(self.line,data_only=True,keep_vba=True)
        ws = wb.active
        sayac=0
        for i in range(7,67,2):
            for j in range(0,8):
                self.row.emit(sayac)
                self.col.emit(0+j)
                print(ws.cell(i,3+j).value)
                self.value.emit(str(ws.cell(i,3+j).value))
                self.writeSignal.emit()
            self.row.emit(sayac)
            self.col.emit(9)
            self.value.emit(str(ws.cell(i,13).value))
            self.writeSignal.emit()
            sayac=sayac+1

        
class GeneralDetailPage(Ui_MainWindow,QtWidgets.QMainWindow):
    def __init__(self,line):
        super().__init__()
        self.setupUi(self)
        self.line=line

        self.row=0
        self.col=0
        self.value=""
        
        self.setTable()

    def setTable(self):
        self.tableWidget.setRowCount(31) 

        self.my_thread = QThread(parent=self)
        self.worker = Worker(self.line)
        self.worker.moveToThread(self.my_thread)

        self.worker.row.connect(self.setRow)
        self.worker.col.connect(self.setCol)
        self.worker.value.connect(self.setValue)
        self.worker.writeSignal.connect(self.write)

        self.my_thread.start()
        self.my_thread.started.connect(self.worker.setTable)
        
    def setRow(self,rowSignal):
        self.row=rowSignal
    
    def setCol(self,colSignal):
        self.col=colSignal
    
    def setValue(self,valueSignal):
        self.value=valueSignal
    
    def write(self):
        self.tableWidget.setItem(self.row,self.col,QTableWidgetItem(self.value))
