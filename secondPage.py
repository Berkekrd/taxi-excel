from PyQt5 import QtWidgets
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtWidgets import * 
from personal import Ui_MainWindow
from openpyxl import load_workbook

class Worker(QObject):

    sayacSignal = pyqtSignal(int)
    kasseGesamt = pyqtSignal(float)
    barSignal = pyqtSignal(str)
    writeSignal = pyqtSignal ()
    finish=pyqtSignal()

    def __init__(self,line,personalRow):
        super().__init__()
        self.line=line
        self.personalRow=personalRow

    def setTableWorker(self):
        wb = load_workbook(self.line)
        ws = wb.active
        sayac=0
        for j in range(7,68,2):
            self.sayacSignal.emit(sayac)
            cValueText1=ws.cell(j,self.personalRow).value
            # cValueText1=cValueText1.replace(",",".")
            self.kasseGesamt.emit(float(cValueText1))
            cValueText2=ws.cell(j,self.personalRow+3).value
            # cValueText2=cValueText2.replace(",",".")
            self.barSignal.emit(str(cValueText2))
            self.writeSignal.emit()
            sayac=sayac+1
        self.finish.emit()

class SecondPage(Ui_MainWindow,QtWidgets.QMainWindow):
    def __init__(self,line,row,personType):
        super().__init__()
        self.setupUi(self)
        self.line=line
        self.personalRow=row
        self.personType=0
        self.sayacTable=0
        self.kasseGesamt = 0
        self.bar=""

        self.setTableTitle()
        self.setTable()

        self.updateButton.clicked.connect(self.excelUpdate)
        self.tableUpdateButton.clicked.connect(self.tableUpdate)

    def setTable(self): # kişinin bir aylık çalışma planını veriyor
        # try:
        #     self.setTableTitle()
        #     wb = load_workbook(self.line)
        #     ws = wb.active
        #     sayac=0
        #     for j in range(7,68,2):
        #         self.tableWidget.setItem(sayac,0,QTableWidgetItem(str(ws.cell(j,self.personalRow).value)))
        #         self.tableWidget.setItem(sayac,1,QTableWidgetItem(str(float(ws.cell(j,self.personalRow).value)*(0.25))))
        #         self.tableWidget.setItem(sayac,2,QTableWidgetItem(str(float(ws.cell(j,self.personalRow).value)*(0.15))))
        #         self.tableWidget.setItem(sayac,3,QTableWidgetItem(str(ws.cell(j,self.personalRow+3).value)))
        #         sayac+=1
        #         self.tableWidget.setRowCount(sayac+1)
        # except:
        #     pass

        try:
            self.my_thread = QThread(parent=self)
            self.worker = Worker(self.line,self.personalRow)
            self.worker.moveToThread(self.my_thread)
            self.worker.sayacSignal.connect(self.setSayac)
            self.worker.barSignal.connect(self.setBar)
            self.worker.kasseGesamt.connect(self.setKasseGesamt)
            self.worker.writeSignal.connect(self.tableWrite)
            self.worker.finish.connect(self.display)
            self.my_thread.start()
            self.my_thread.started.connect(self.worker.setTableWorker)
        except Exception as e:
            self.errorLine.setText(f"Thread Problemi. {e}")


    def setBar(self,text):
        self.bar=text

    def setKasseGesamt(self,text):
        self.kasseGesamt=text

    def setSayac(self,text):
        self.sayacTable=text

    def tableWrite(self):
        try:
            if self.personType==0:
                self.tableWidget.setItem(self.sayacTable,0,QTableWidgetItem(str(self.kasseGesamt)))
                self.tableWidget.setItem(self.sayacTable,1,QTableWidgetItem(str(float(self.kasseGesamt)*(0.15))))
                self.tableWidget.setItem(self.sayacTable,2,QTableWidgetItem(str(float(self.kasseGesamt)*(0.25))))
                self.tableWidget.setItem(self.sayacTable,3,QTableWidgetItem(str(self.bar)))
        except Exception as e:
            self.errorLine.setText(f"Tabloya veri yazılamadı. {e}")

        
    def setTableTitle(self):
        self.tableWidget.setRowCount(31)
        try:
            if self.personType==0:
                self.tableWidget.setHorizontalHeaderItem(1, QTableWidgetItem("15%"))
                self.tableWidget.setHorizontalHeaderItem(2, QTableWidgetItem("25%"))
                self.tableWidget.setHorizontalHeaderItem(3, QTableWidgetItem("BAR"))
            elif self.personType==1:
                pass
            else:
                self.errorLine.setText("Hatalı işlem Yapıldı.")
        except Exception as e:
            self.errorLine.setText(f"Tablo düzenlenemedi. {e}")

    def display(self):
        self.errorLine.setText("işlem tamamlandı")

    def messageBox(self,displayText,boxTitle):
        msgBox = QMessageBox()
        msgBox.about(self,boxTitle,displayText)

    def excelUpdate(self):
        try:
            wb = load_workbook(self.line)
            ws = wb.active
            sayac=0
            for j in range(7,68,2):
                c1 = ws.cell(row = j, column = self.personalRow)
                cValueText1=self.tableWidget.item(sayac,0).text()
                # cValueText1=cValueText1.replace(".",",")
                c1.value=float(cValueText1)
                c2 = ws.cell(row = j, column = self.personalRow+3)
                cValueText2=self.tableWidget.item(sayac,3).text()
                # cValueText2=cValueText2.replace(".",",")
                c2.value=float(cValueText2)
                sayac=sayac+1
            wb.save(self.line)
            self.messageBox("EXCEL GÜNCELLENDİ","EXCEL")
        except Exception as e:
            self.errorLine.setText(f"Excel Güncellemesi Yapılamadı.{e}")
    

    def tableUpdate(self):
        try:
            for i in range (0,31):
                tableValue=float(self.tableWidget.item(i,0).text())
                self.tableWidget.setItem(i,1,QTableWidgetItem(str(tableValue*(0.15))))
                self.tableWidget.setItem(i,2,QTableWidgetItem(str(tableValue*(0.25))))
        except Exception as e:
            self.errorLine.setText(f"Tablo güncellenmedi. {e}")